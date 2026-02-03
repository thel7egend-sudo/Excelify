from PySide6.QtWidgets import QMainWindow, QWidget, QVBoxLayout, QFileDialog, QMessageBox
from top_chrome import TopChrome
from home_page import HomePage
from editor_page import EditorPage
from storage import save_state, load_state
from document import Document
from storage import save_state, load_state
from PySide6.QtWidgets import QFileDialog, QMessageBox
from openpyxl import Workbook, load_workbook
from PySide6.QtGui import QPalette, QColor
from PySide6.QtWidgets import QApplication



class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.resize(1200, 800)
        self.setWindowTitle("Excelify")

        self.chrome = TopChrome()
        self.chrome.home_clicked.connect(self.go_home)

        self.container = QWidget()
        self.container_layout = QVBoxLayout(self.container)
        self.container_layout.setContentsMargins(0, 0, 0, 0)

        self.home = HomePage(self.chrome)
        self.home.open_import_requested.connect(self.import_excel)
        self.home.toggle_dark_mode_requested.connect(self.toggle_dark_mode)

        self.home.open_document_requested.connect(self.open_editor_for_document)
        self.home.save_requested.connect(self.save_app_state)

        self.load_app_state()
        
        self.editor = None

        self.container_layout.addWidget(self.chrome)
        self.container_layout.addWidget(self.home)

        self.setCentralWidget(self.container)
        self.is_grid_dark = False




    def go_home(self):
        if self.editor:
            self.editor.hide()

        self.chrome.show_home_mode()
        self.home.show()

    def open_editor_for_document(self, document):
        if self.editor:
            self.container_layout.removeWidget(self.editor)
            self.editor.deleteLater()

        self.editor = EditorPage(document)
        # apply grid-only dark mode if enabled
        self.editor.apply_grid_dark_mode(self.is_grid_dark)

        # ✅ CONNECT SAVE HERE (parent is now MainWindow)
        self.editor.model.save_requested.connect(self.save_app_state)
        self.editor.document_changed.connect(self.save_app_state)
        self.editor.export_requested.connect(self.export_document_to_excel)

        self.container_layout.addWidget(self.editor)
        self.home.hide()
        self.chrome.show_editor_mode()
        # apply grid dark mode if enabled

    def save_app_state(self):
        state = {
            "documents": [doc.to_dict() for doc in self.home.documents]
        }
        save_state(state)

    def load_app_state(self):
        state = load_state()
        if not state:
            return

        for doc_data in state.get("documents", []):
            doc = Document.from_dict(doc_data)
            self.home.documents.append(doc)
            self.home.add_existing_document(doc)

    def load_app_state(self):
        state = load_state()
        if not state:
            return

        for doc_data in state.get("documents", []):
            doc = Document.from_dict(doc_data)
            self.home.documents.append(doc)
            self.home.add_existing_document(doc)
    def export_document_to_excel(self, document):
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Export to Excel",
            f"{document.name}.xlsx",
            "Excel Files (*.xlsx)"
        )

        if not path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = document.name[:31]  # Excel limit

        # determine grid bounds
    def export_document_to_excel(self, document):
        from openpyxl import Workbook
        from PySide6.QtWidgets import QFileDialog, QMessageBox

        path, _ = QFileDialog.getSaveFileName(
            self,
            "Export to Excel",
            f"{document.name}.xlsx",
            "Excel Files (*.xlsx)"
        )

        if not path:
            return

        wb = Workbook()

        # remove default sheet
        default_ws = wb.active
        wb.remove(default_ws)

        has_data = False

        for sheet in document.sheets:
            ws = wb.create_sheet(title=sheet.name[:31])  # Excel limit

            if not sheet.cells:
                continue

            has_data = True
            used_cells = sheet.cells.keys()

            max_row = max(r for r, _ in used_cells)
            max_col = max(c for _, c in used_cells)

            for row in range(max_row + 1):
                for col in range(max_col + 1):
                    value = sheet.cells.get((row, col), "")
                    ws.cell(row=row + 1, column=col + 1, value=value)

        if not has_data:
            QMessageBox.information(
                self,
                "Nothing to Export",
                "This document has no data."
            )
            return

        wb.save(path)

        QMessageBox.information(
            self,
            "Export Complete",
            "Excel file exported successfully."
        )
    def import_excel(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Import Excel File",
            "",
            "Excel Files (*.xlsx)"
        )

        if not path:
            return

        if not path.lower().endswith(".xlsx"):
            QMessageBox.warning(
                self,
                "Invalid File",
                "Only .xlsx files are supported."
            )
            return

        try:
            wb = load_workbook(path, data_only=True)
        except Exception as e:
            QMessageBox.critical(
                self,
                "Import Failed",
                f"Could not open file:\n{e}"
            )
            return

        # create new document
        from document import Document, Sheet

        doc_name = path.split("/")[-1].rsplit(".", 1)[0]
        document = Document(doc_name)
        document.sheets.clear()

        for ws in wb.worksheets:
            sheet = Sheet(ws.title)

            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        r = cell.row - 1
                        c = cell.column - 1
                        sheet.cells[(r, c)] = str(cell.value)

            document.sheets.append(sheet)

        if not document.sheets:
            document.sheets.append(Sheet("Sheet1"))

        document.active_sheet_index = 0

        # add to home page
        self.home.documents.append(document)
        self.home.add_existing_document(document)
        self.save_app_state()

    def toggle_dark_mode(self):
        self.is_grid_dark = not self.is_grid_dark

        # grid / editor
        if self.editor:
            self.editor.apply_grid_dark_mode(self.is_grid_dark)

        # top bar (HOME area)
        self.chrome.apply_dark_mode(self.is_grid_dark)

